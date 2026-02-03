"""File content extraction and query injection for ResearchBot."""

import csv
import logging
import sqlite3
from pathlib import Path
from typing import List

from PyPDF2 import PdfReader
from docx import Document
from openpyxl import load_workbook

from utils.models import UploadedFile

logger = logging.getLogger(__name__)


class FileContextInjector:
    """Handles file content extraction and injection into queries."""

    @staticmethod
    def extract_file_content(file_path: str) -> str:
        """Extract text content from a file based on its type."""
        path = Path(file_path)

        if not path.exists():
            logger.error(f"File not found: {file_path}")
            raise FileNotFoundError(f"File not found: {file_path}")

        suffix = path.suffix.lower()

        # Text-based file types that can be read directly
        text_types = {
            ".txt", ".md", ".markdown", ".json", ".xml", ".yaml", ".yml",
            ".html", ".htm", ".rtf", ".py", ".js", ".ts", ".jsx", ".tsx",
            ".java", ".c", ".cpp", ".h", ".hpp", ".css", ".scss", ".sass",
            ".sql", ".sh", ".bash", ".go", ".rs", ".rb", ".php",
            ".swift", ".kt", ".scala", ".r", ".ipynb",
            ".log", ".ini", ".conf", ".cfg", ".env", ".gitignore", ".dockerignore"
        }

        extractors = {
            ".pdf": FileContextInjector._extract_pdf,
            ".docx": FileContextInjector._extract_docx,
            ".csv": FileContextInjector._extract_csv,
            ".xlsx": FileContextInjector._extract_xlsx,
            ".xls": FileContextInjector._extract_xlsx,
            ".tsv": FileContextInjector._extract_tsv,
            ".sqlite": FileContextInjector._extract_sqlite,
            ".sqlite3": FileContextInjector._extract_sqlite,
            ".db": FileContextInjector._extract_sqlite,
        }

        # Add all text types to use txt extractor
        for ext in text_types:
            extractors[ext] = FileContextInjector._extract_txt

        if suffix not in extractors:
            # Try to read as text anyway
            logger.warning(f"Unknown file type {suffix}, attempting to read as text")
            extractors[suffix] = FileContextInjector._extract_txt

        content = extractors[suffix](file_path)
        logger.info(f"Extracted {len(content)} characters from {path.name}")
        return content

    @staticmethod
    def _extract_pdf(file_path: str) -> str:
        """Extract text from a PDF file."""
        try:
            with open(file_path, "rb") as file:
                reader = PdfReader(file)
                text_parts = []
                total_pages = len(reader.pages)
                logger.info(f"PDF has {total_pages} pages")

                for page_num, page in enumerate(reader.pages):
                    try:
                        page_text = page.extract_text()
                        if page_text and page_text.strip():
                            text_parts.append(page_text)
                            logger.debug(f"Page {page_num + 1}: extracted {len(page_text)} chars")
                        else:
                            logger.warning(f"Page {page_num + 1}: no text extracted (may be scanned/image)")
                    except Exception as page_err:
                        logger.warning(f"Page {page_num + 1} extraction failed: {page_err}")

                result = "\n\n".join(text_parts).strip()

                if not result:
                    logger.warning(f"PDF extraction returned empty - file may be scanned or image-based")
                    return "[PDF appears to be scanned or image-based - no text could be extracted]"

                logger.info(f"Extracted {len(result)} total characters from PDF")
                return result
        except Exception as e:
            logger.error(f"Error extracting PDF: {e}")
            raise ValueError(f"Failed to extract PDF content: {e}")

    @staticmethod
    def _extract_docx(file_path: str) -> str:
        """Extract text from a DOCX file."""
        try:
            doc = Document(file_path)
            paragraphs = [para.text for para in doc.paragraphs if para.text.strip()]
            return "\n\n".join(paragraphs).strip()
        except Exception as e:
            logger.error(f"Error extracting DOCX: {e}")
            raise ValueError(f"Failed to extract DOCX content: {e}")

    @staticmethod
    def _extract_txt(file_path: str) -> str:
        """Extract text from a plain text file."""
        try:
            with open(file_path, "r", encoding="utf-8") as file:
                return file.read().strip()
        except UnicodeDecodeError:
            with open(file_path, "r", encoding="latin-1") as file:
                return file.read().strip()
        except Exception as e:
            logger.error(f"Error extracting TXT: {e}")
            raise ValueError(f"Failed to extract TXT content: {e}")

    @staticmethod
    def _extract_csv(file_path: str) -> str:
        """Extract CSV content and format as markdown table."""
        try:
            with open(file_path, "r", encoding="utf-8") as file:
                reader = csv.reader(file)
                rows = list(reader)

            if not rows:
                return ""

            headers = rows[0]
            header_line = "| " + " | ".join(headers) + " |"
            separator = "|" + "|".join(["---"] * len(headers)) + "|"

            data_lines = []
            for row in rows[1:]:
                if len(row) == len(headers):
                    data_lines.append("| " + " | ".join(row) + " |")

            return "\n".join([header_line, separator] + data_lines)

        except Exception as e:
            logger.error(f"Error extracting CSV: {e}")
            raise ValueError(f"Failed to extract CSV content: {e}")

    @staticmethod
    def _extract_xlsx(file_path: str) -> str:
        """Extract Excel content and format as markdown table."""
        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            all_sheets = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                rows = list(sheet.iter_rows(values_only=True))

                if not rows:
                    continue

                # Filter out completely empty rows
                rows = [r for r in rows if any(cell is not None for cell in r)]
                if not rows:
                    continue

                # Convert None to empty string and all values to strings
                rows = [[str(cell) if cell is not None else "" for cell in row] for row in rows]

                headers = rows[0]
                header_line = "| " + " | ".join(headers) + " |"
                separator = "|" + "|".join(["---"] * len(headers)) + "|"

                data_lines = []
                for row in rows[1:]:
                    # Pad row if shorter than headers
                    padded_row = row + [""] * (len(headers) - len(row))
                    data_lines.append("| " + " | ".join(padded_row[:len(headers)]) + " |")

                sheet_content = f"**Sheet: {sheet_name}**\n\n" + "\n".join([header_line, separator] + data_lines)
                all_sheets.append(sheet_content)

            workbook.close()
            return "\n\n".join(all_sheets)

        except Exception as e:
            logger.error(f"Error extracting Excel: {e}")
            raise ValueError(f"Failed to extract Excel content: {e}")

    @staticmethod
    def _extract_tsv(file_path: str) -> str:
        """Extract TSV content and format as markdown table."""
        try:
            with open(file_path, "r", encoding="utf-8") as file:
                reader = csv.reader(file, delimiter="\t")
                rows = list(reader)

            if not rows:
                return ""

            headers = rows[0]
            header_line = "| " + " | ".join(headers) + " |"
            separator = "|" + "|".join(["---"] * len(headers)) + "|"

            data_lines = []
            for row in rows[1:]:
                if len(row) == len(headers):
                    data_lines.append("| " + " | ".join(row) + " |")

            return "\n".join([header_line, separator] + data_lines)

        except Exception as e:
            logger.error(f"Error extracting TSV: {e}")
            raise ValueError(f"Failed to extract TSV content: {e}")

    @staticmethod
    def _extract_sqlite(file_path: str) -> str:
        """Extract SQLite database schema and sample data."""
        try:
            conn = sqlite3.connect(file_path)
            cursor = conn.cursor()

            # Get all table names
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
            tables = [row[0] for row in cursor.fetchall()]

            if not tables:
                conn.close()
                return "[Empty database - no tables found]"

            all_tables = []
            for table_name in tables:
                # Get column info
                cursor.execute(f"PRAGMA table_info({table_name})")
                columns = cursor.fetchall()
                col_names = [col[1] for col in columns]

                # Get row count
                cursor.execute(f"SELECT COUNT(*) FROM {table_name}")
                row_count = cursor.fetchone()[0]

                # Get sample data (first 10 rows)
                cursor.execute(f"SELECT * FROM {table_name} LIMIT 10")
                sample_rows = cursor.fetchall()

                # Build markdown table
                header_line = "| " + " | ".join(col_names) + " |"
                separator = "|" + "|".join(["---"] * len(col_names)) + "|"

                data_lines = []
                for row in sample_rows:
                    row_str = [str(cell) if cell is not None else "" for cell in row]
                    data_lines.append("| " + " | ".join(row_str) + " |")

                table_content = f"**Table: {table_name}** ({row_count} rows)\n\n"
                table_content += "\n".join([header_line, separator] + data_lines)
                all_tables.append(table_content)

            conn.close()
            return "\n\n".join(all_tables)

        except Exception as e:
            logger.error(f"Error extracting SQLite: {e}")
            raise ValueError(f"Failed to extract SQLite content: {e}")

    @staticmethod
    def build_file_context(files: List[UploadedFile]) -> str:
        """Build combined context from multiple uploaded files."""
        if not files:
            return ""

        context_parts = ["## UPLOADED FILE CONTEXT\n"]

        for file in files:
            try:
                content = FileContextInjector.extract_file_content(file.path)
                context_parts.append(f"### File: {file.filename}")
                context_parts.append(content)
                context_parts.append("")
            except Exception as e:
                logger.warning(f"Skipping file {file.filename}: {e}")
                context_parts.append(f"### File: {file.filename}")
                context_parts.append(f"[Error extracting content: {e}]")
                context_parts.append("")

        return "\n".join(context_parts)

    @staticmethod
    def inject_into_query(query_text: str, file_context: str) -> str:
        """Inject file context into the query prompt."""
        if not file_context:
            return query_text

        return f"{file_context}\n\n## QUERY\n\n{query_text}"
